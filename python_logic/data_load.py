from openpyxl import load_workbook,Workbook
import psycopg2
from openpyxl.styles import PatternFill


def write_data_to_postgres(columns_name, data, db_name, user, password, host, port):

    conn = psycopg2.connect(database=db_name, user=user, password=password, host=host, port=port)
    cursor = conn.cursor()

    try:
        cursor.execute(f"""CREATE TABLE IF NOT EXISTS public.overdue ("{columns_name[0]}" TEXT, "{columns_name[1]}" TEXT,
                       "{columns_name[2]}" BIGINT, "{columns_name[3]}" TEXT,"{columns_name[4]}" TEXT,"{columns_name[5]}" BIGINT,
                       "{columns_name[6]}" TEXT,"{columns_name[7]}" BIGINT,"{columns_name[8]}" BIGINT,"{columns_name[9]}" BIGINT,
                       "{columns_name[10]}" DATE, "{columns_name[11]}" BIGINT)""")

        for row in data:
            cursor.execute(f"""INSERT INTO public.overdue ("{columns_name[0]}", "{columns_name[1]}", "{columns_name[2]}",
                           "{columns_name[3]}","{columns_name[4]}","{columns_name[5]}","{columns_name[6]}",
                           "{columns_name[7]}","{columns_name[8]}","{columns_name[9]}","{columns_name[10]}",
                           "{columns_name[11]}") VALUES ('{row[0]}','{row[1]}', {row[2]},'{row[3]}','{row[4]}',
                           {row[5]},'{row[6]}',{row[7]},{row[8]},{row[9]}, TO_DATE('{row[10].replace('.','-')}','DD/MM/YYYY')
                           ,{row[11]})  """)

        conn.commit()

        print("Данные успешно записаны в базу данных.")
    except (Exception, psycopg2.Error) as error:
        print("Ошибка при записи данных в базу данных:", error)
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()


def read_postgres_to_tg(db_name, user, password, host, port):
    conn = psycopg2.connect(database=db_name, user=user, password=password, host=host, port=port)
    cursor = conn.cursor()

    try:
        cursor.execute(f"""SELECT "Субъект РФ", SUM("Количество\nДоз"), ROUND(AVG("Просрочено дней"),0) from public.overdue
                           GROUP BY "Субъект РФ" """)
        conn.commit()
        data_postgres = cursor.fetchall()
        workbook = Workbook()
        sheet = workbook.active
        red_fill = PatternFill(start_color='FFCCCC00',
                              end_color='FFCCCC00',
                              fill_type='solid')
        orange_fill = PatternFill(start_color='F4B18300',
                              end_color='F4B18300',
                              fill_type='solid')
        yellow_fill = PatternFill(start_color='FFFFCC00',
                              end_color='FFFFCC00',
                              fill_type='solid')

        sheet['A1'] = 'Субъект РФ'
        sheet['B1'] = 'Количество\nДоз'
        sheet['C1'] = 'Просрочено дней'

        i = 2
        for data in data_postgres:
            sheet[f'A{i}'] = data[0]
            sheet[f'B{i}'] = data[1]
            sheet[f'C{i}'] = data[2]
            if data[2] >= 1 and data[2] <= 7:
                sheet[f'C{i}'].fill = yellow_fill
            elif data[2] >= 8 and data[2] <=14:
                sheet[f'C{i}'].fill = orange_fill
            elif data[2] > 14:
                sheet[f'C{i}'].fill = red_fill
            i +=1



        workbook.save('answer.xlsx')

        print("Данные успешно записаны в excel.")
    except (Exception, psycopg2.Error) as error:
        print("Ошибка при выгрузки данных:", error)
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()


def read_data(path_data):
    data_wb = load_workbook(path_data)
    sheet = data_wb['Статика']

    data = []
    columns_name = []

    for row in sheet.iter_rows(min_row=4,values_only=True,max_row=4,max_col=12):
        columns_name = row

    for row in sheet.iter_rows(min_row=6,values_only=True,max_col=12):
        data.append(row)

    return columns_name, data

